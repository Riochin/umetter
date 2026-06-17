from pathlib import Path
import openpyxl

BASE_DIR = Path(__file__).resolve().parents[1]
XLSX_PATH = BASE_DIR / "data" / "raw" / "tsuda_gakugei_timetable_2026.xlsx"

def normalize(value):
    if value is None:
        return ""
    return str(value).replace("\n", " ").strip()

def main():
    if not XLSX_PATH.exists():
        raise FileNotFoundError(f"{XLSX_PATH} がありません。先にcurlで取得してください。")

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    print("=== シート一覧 ===")
    for i, sheet_name in enumerate(wb.sheetnames, start=1):
        print(f"{i}: {sheet_name}")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        print("\n" + "=" * 80)
        print(f"シート名: {sheet_name}")
        print(f"最大行: {ws.max_row}, 最大列: {ws.max_column}")
        print("=" * 80)

        # 最初の40行・20列だけ表示して、表の形を見る
        for row_idx in range(1, min(ws.max_row, 40) + 1):
            values = []
            for col_idx in range(1, min(ws.max_column, 20) + 1):
                value = normalize(ws.cell(row=row_idx, column=col_idx).value)
                values.append(value)

            # 全部空の行は飛ばす
            if any(values):
                print(f"{row_idx:03d}: " + " | ".join(values))

if __name__ == "__main__":
    main()
