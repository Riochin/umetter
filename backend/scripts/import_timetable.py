from pathlib import Path
import re
import sqlite3
import urllib.request

import openpyxl


BASE_DIR = Path(__file__).resolve().parents[1]
DB_PATH = BASE_DIR / "umetter.db"
DATA_DIR = BASE_DIR / "data" / "raw"
XLSX_PATH = DATA_DIR / "tsuda_gakugei_timetable_2026.xlsx"

TIMETABLE_URL = (
    "https://docs.google.com/spreadsheets/d/e/"
    "2PACX-1vR45LvUqliCyhWy5Kq8L3IShriPz4lSGMuftbOnBv5MKRwOujDc19zhiqFuJP3v_XKsfAf3oO9IMrFS/"
    "pub?output=xlsx"
)

DAY_MAP = {
    "月": 1,
    "火": 2,
    "水": 3,
    "木": 4,
    "金": 5,
    "土": 6,
}


def clean_text(value) -> str:
    if value is None:
        return ""

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    text = str(value).replace("\n", " ").strip()
    text = re.sub(r"\s+", " ", text)
    text = text.removeprefix("★").strip()

    if re.fullmatch(r"\d+\.0", text):
        return text[:-2]

    return text


def to_period(value) -> int | None:
    if value is None or value == "":
        return None

    try:
        return int(float(value))
    except ValueError:
        return None


def expand_terms(term_text: str) -> list[str]:
    """
    担当期の表記を T1/T2/T3/T4 単位に展開する。

    対応例:
    - T1      -> ["T1"]
    - T13     -> ["T1", "T3"]
    - T1234   -> ["T1", "T2", "T3", "T4"]
    - T1/T3   -> ["T1", "T3"]
    - T1・3   -> ["T1", "T3"]
    - T1-4    -> ["T1", "T2", "T3", "T4"]
    """
    text = clean_text(term_text).upper()

    if not text:
        return []

    terms = set()

    # T1-4 / T1~T4 / 1-4 のような範囲表記に対応
    for start, end in re.findall(r"T?\s*([1-4])\s*[-~〜]\s*T?\s*([1-4])", text):
        start_num = int(start)
        end_num = int(end)

        low = min(start_num, end_num)
        high = max(start_num, end_num)

        for num in range(low, high + 1):
            terms.add(f"T{num}")

    # T13 / T1234 / T1/T3 / T1・3 / T1,T3 などに対応
    for group in re.findall(r"T\s*((?:[1-4]\s*(?:[/,，、・･+&]\s*)?)+)", text):
        for digit in re.findall(r"[1-4]", group):
            terms.add(f"T{digit}")

    # 念のため、Tなしで 13 / 1234 のように書かれている場合にも対応
    compact_text = re.sub(r"[\s/,，、・･+&]+", "", text)
    if not terms and re.fullmatch(r"[1-4]+", compact_text):
        for digit in compact_text:
            terms.add(f"T{digit}")

    return sorted(terms, key=lambda term: int(term[1]))


def make_class_id(course_code: str, term: str, day_of_week: int, period: int) -> str:
    safe_code = re.sub(r"[^A-Za-z0-9_-]", "_", course_code)
    safe_term = re.sub(r"[^A-Za-z0-9_-]", "_", term)
    return f"tt_{safe_code}_{safe_term}_{day_of_week}_{period}"


def download_xlsx() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    if XLSX_PATH.exists():
        print(f"既存ファイルを使います: {XLSX_PATH}")
        return

    print("時間割ファイルをダウンロードします...")
    urllib.request.urlretrieve(TIMETABLE_URL, XLSX_PATH)
    print(f"保存しました: {XLSX_PATH}")


def parse_t134(ws):
    """
    T134:
    6行目: ヘッダー
    A:曜日 B:時限 C:時間割コード D:授業名 E:開講期 F:担当期 G:教員名 H:教室名称

    T134シートには T1/T3/T4 のほか、T13/T1234 のような複合ターム表記もあり得るため、
    担当期を T1/T2/T3/T4 単位に展開して保存する。
    """
    rows = []

    for row in ws.iter_rows(min_row=7, values_only=True):
        day = clean_text(row[0])
        period = to_period(row[1])
        course_code = clean_text(row[2])
        name = clean_text(row[3])
        term_text = clean_text(row[5])
        terms = expand_terms(term_text)
        teacher_name = clean_text(row[6])
        room = clean_text(row[7])

        if not day or period is None or not course_code or not name or not terms:
            continue

        day_of_week = DAY_MAP.get(day)
        if day_of_week is None:
            continue

        for term in terms:
            rows.append({
                "id": make_class_id(course_code, term, day_of_week, period),
                "course_code": course_code,
                "term": term,
                "name": name,
                "teacher_name": teacher_name,
                "day_of_week": day_of_week,
                "period": period,
                "room": room,
            })

    return rows


def parse_t2(ws):
    """
    T2:
    7行目: ヘッダー
    A:時間割コード B:授業名 C:教員氏名 D:教室名称 E:曜日 F:時限 G:授業年月日

    T2シートは第2ターム用なので、termはT2固定にする。
    """
    rows = []

    for row in ws.iter_rows(min_row=9, values_only=True):
        course_code = clean_text(row[0])
        name = clean_text(row[1])
        teacher_name = clean_text(row[2])
        room = clean_text(row[3])
        day = clean_text(row[4])
        period = to_period(row[5])
        term = "T2"

        if not day or period is None or not course_code or not name:
            continue

        day_of_week = DAY_MAP.get(day)
        if day_of_week is None:
            continue

        rows.append({
            "id": make_class_id(course_code, term, day_of_week, period),
            "course_code": course_code,
            "term": term,
            "name": name,
            "teacher_name": teacher_name,
            "day_of_week": day_of_week,
            "period": period,
            "room": room,
        })

    return rows


def dedupe_classes(classes):
    unique = {}

    for cls in classes:
        if cls["id"] not in unique:
            unique[cls["id"]] = cls

    return list(unique.values())


def ensure_columns(conn):
    columns = {
        row[1]
        for row in conn.execute("PRAGMA table_info(classes)").fetchall()
    }

    if "course_code" not in columns:
        conn.execute("ALTER TABLE classes ADD COLUMN course_code TEXT NOT NULL DEFAULT ''")

    if "term" not in columns:
        conn.execute("ALTER TABLE classes ADD COLUMN term TEXT NOT NULL DEFAULT ''")


def import_classes(classes):
    conn = sqlite3.connect(DB_PATH)

    try:
        conn.execute("PRAGMA foreign_keys = ON")
        ensure_columns(conn)

        sql = """
        INSERT INTO classes (
            id,
            course_code,
            term,
            name,
            teacher_name,
            day_of_week,
            period,
            room,
            is_canceled
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, 0)
        ON CONFLICT(id) DO UPDATE SET
            course_code = excluded.course_code,
            term = excluded.term,
            name = excluded.name,
            teacher_name = excluded.teacher_name,
            day_of_week = excluded.day_of_week,
            period = excluded.period,
            room = excluded.room;
        """

        conn.executemany(
            sql,
            [
                (
                    cls["id"],
                    cls["course_code"],
                    cls["term"],
                    cls["name"],
                    cls["teacher_name"],
                    cls["day_of_week"],
                    cls["period"],
                    cls["room"],
                )
                for cls in classes
            ],
        )

        conn.commit()

    finally:
        conn.close()


def main():
    download_xlsx()

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    all_classes = []

    if "T134" in wb.sheetnames:
        all_classes.extend(parse_t134(wb["T134"]))

    if "T2" in wb.sheetnames:
        all_classes.extend(parse_t2(wb["T2"]))

    unique_classes = dedupe_classes(all_classes)

    print(f"抽出件数: {len(all_classes)}")
    print(f"重複排除後: {len(unique_classes)}")

    print("\nサンプル:")
    for cls in unique_classes[:10]:
        print(cls)

    import_classes(unique_classes)

    print("\nclasses テーブルへの流し込みが完了しました。")


if __name__ == "__main__":
    main()
